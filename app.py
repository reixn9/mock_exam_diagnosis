from flask import Flask, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from werkzeug.utils import secure_filename
import os
import re

app = Flask(__name__)
app.secret_key = "mock-exam-secret"

# 캐시 최소화 (항상 최신 템플릿/정적파일 사용)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ---------- 유틸 함수 ----------

def parse_answers(text: str):
    """
    공백/줄바꿈/콤마/붙여쓰기 모두 허용하고,
    1~5인 숫자만 int로 반환 (엑셀에서 녹색 삼각형 방지).
    """
    if not text:
        return []
    tokens = re.split(r"[\s,]+", text.strip())
    tokens = [t for t in tokens if t]

    joined = "".join(tokens)
    out = []
    if joined and all(ch in "12345" for ch in joined):
        out = [int(ch) for ch in joined]
    else:
        for t in tokens:
            if t.isdigit():
                v = int(t)
                if 1 <= v <= 5:
                    out.append(v)
    return out


def validate_range(label, start, end, answers_len, errors):
    """문항 범위와 답 개수 기본 검증"""
    if start is None or end is None:
        errors.append(f"{label} 범위를 입력하세요.")
        return
    if not (1 <= start <= 45 and 1 <= end <= 45) or start > end:
        errors.append(f"{label} 범위가 올바르지 않습니다. (1~45, 시작<=끝)")
        return
    expected = end - start + 1
    if answers_len != expected:
        errors.append(f"{label} 답 개수는 {expected}개여야 합니다. (지금 {answers_len}개)")


def write_correct_answers(ws, start, end, answers):
    """정답시트: A열에 1~45, B열에 지정 범위만 정답 기록"""
    for i in range(1, 46):
        ws.cell(row=i + 1, column=1, value=i)
    for offset, q in enumerate(range(start, end + 1), start=0):
        ws.cell(row=q + 1, column=2, value=int(answers[offset]))


def append_or_update_student(ws, student_name, student_answers, start, end):
    """
    같은 이름이면 같은 열에 이어서 덮어쓰기,
    없으면 새 열 생성. 지정 범위만 기록, 오답 노란색.
    """
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 기존 열 탐색
    target_col = None
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == student_name:
            target_col = col
            break

    # 없으면 새 열 생성
    if target_col is None:
        target_col = ws.max_column + 1
        ws.cell(row=1, column=target_col, value=student_name)

    # 지정 범위만 덮어쓰기
    for offset, q in enumerate(range(start, end + 1), start=0):
        val = int(student_answers[offset])
        cell = ws.cell(row=q + 1, column=target_col, value=val)
        correct = ws.cell(row=q + 1, column=2).value
        if correct is not None and val != correct:
            cell.fill = yellow


def build_new_workbook(grade, year, month, level, round_no, school_name, cat_payload: dict):
    """
    새 엑셀 생성. cat_payload[cat] = {start, end, answers}
    """
    wb = Workbook()
    wb.remove(wb.active)

    for cat, payload in cat_payload.items():
        ws = wb.create_sheet(title=cat)
        ws["A1"] = "문항 번호"
        ws["B1"] = "정답"
        write_correct_answers(ws, payload["start"], payload["end"], payload["answers"])

    # 파일명 규칙
    if level == "학교" and school_name:
        filename = f"학교({school_name}) {round_no}회 {year}학년도 {month}월 고{grade} 모의고사 진단지.xlsx"
    else:
        filename = f"{level} {round_no}회 {year}학년도 {month}월 고{grade} 모의고사 진단지.xlsx"

    wb.save(filename)
    return filename


# ---------- 공통 설정 ----------

@app.after_request
def no_cache(resp):
    if resp.mimetype in ("text/html", "application/javascript", "text/css", "application/json"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp


@app.route("/_version")
def _version():
    # 버전 문자열로 실제 배포 코드 확인용
    return "VERSION: 2025-11-10-1"


# ---------- 메인 라우트 ----------

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        mode = request.form.get("mode", "none")

        # ======================
        # 1) 새 파일 만들기
        # ======================
        if mode == "new":
            form = request.form
            errors = []

            grade = form.get("grade", "").strip()
            year = form.get("year", "").strip()
            month = form.get("month", "").strip()
            level = form.get("level", "").strip()
            school_name = form.get("school_name", "").strip()
            round_no = form.get("round", "").strip()
            selected_cats = form.getlist("categories")
            student_count = int(form.get("new_student_count", 0))

            # 학년/연도/월/회차/레벨 검증
            if grade not in {"1", "2", "3"}:
                errors.append("학년(고1/2/3)을 선택하세요.")
            if not year.isdigit():
                errors.append("학년도를 숫자로 입력하세요.")
            if not month.isdigit():
                errors.append("월을 숫자로 입력하세요.")
            if not round_no.isdigit():
                errors.append("회차를 선택하세요.")
            if level not in {"입문", "기본", "실전", "학교"}:
                errors.append("레벨을 선택하세요.")
            if level == "학교" and not school_name:
                errors.append("학교 레벨을 선택했다면 학교명을 입력하세요.")

            # 카테고리 규칙
            if not selected_cats:
                errors.append("최소 한 개의 카테고리를 선택하세요.")
            if "화작" in selected_cats and "언매" not in selected_cats:
                errors.append("화작을 선택하면 언매 정답도 함께 입력해야 합니다.")
            if "언매" in selected_cats and "화작" not in selected_cats:
                errors.append("언매를 선택하면 화작 정답도 함께 입력해야 합니다.")

            # 정답들
            cat_payload = {}
            for cat in ["공통", "화작", "언매"]:
                if cat in selected_cats:
                    raw = form.get(f"answers_{cat}", "")
                    parsed = parse_answers(raw)
                    try:
                        s = int(form.get(f"answers_{cat}_start", 1))
                        e = int(form.get(f"answers_{cat}_end", 45))
                    except ValueError:
                        s, e = None, None
                    validate_range(f"{cat} 정답", s, e, len(parsed), errors)
                    bad = [x for x in parsed if x not in (1, 2, 3, 4, 5)]
                    if bad:
                        errors.append(f"{cat} 정답에 1~5가 아닌 값이 있습니다: {sorted(set(bad))}")
                    cat_payload[cat] = {"start": s, "end": e, "answers": parsed}

            # 학생들
            students = []
            for i in range(1, student_count + 1):
                name = form.get(f"new_student_{i}_name", "").strip()
                cat = form.get(f"new_student_{i}_category", "").strip()
                raw = form.get(f"new_student_{i}_answers", "")
                try:
                    s = int(form.get(f"new_student_{i}_start", 1))
                    e = int(form.get(f"new_student_{i}_end", 45))
                except ValueError:
                    s, e = None, None

                if not (name or raw):
                    continue  # 완전 빈 행은 스킵

                parsed = parse_answers(raw)
                validate_range(f"{name}({cat})", s, e, len(parsed), errors)
                if cat not in cat_payload:
                    errors.append(f"{name} 학생의 카테고리 '{cat}'에 대한 정답이 없습니다.")
                students.append({"name": name, "category": cat, "answers": parsed, "start": s, "end": e})

            # 에러 있으면: 입력값 유지해서 다시 렌더링
            if errors:
                return render_template(
                    "index.html",
                    error_messages=errors,
                    last_mode="new",
                    form_data=form
                )

            # 엑셀 생성 + 학생들 반영
            filename = build_new_workbook(
                grade=grade,
                year=year,
                month=month,
                level=level,
                round_no=round_no,
                school_name=school_name,
                cat_payload=cat_payload
            )
            wb = load_workbook(filename)
            for s in students:
                ws = wb[s["category"]]
                append_or_update_student(ws, s["name"], s["answers"], s["start"], s["end"])
            wb.save(filename)
            return send_file(filename, as_attachment=True)

        # ======================
        # 2) 기존 파일에 학생 추가
        # ======================
        if mode == "add":
            form = request.form
            uploaded = request.files.get("excel_file")
            student_count = int(form.get("add_student_count", 0))
            errors = []

            if not uploaded or uploaded.filename == "":
                errors.append("엑셀 파일을 업로드하세요.")
                return render_template(
                    "index.html",
                    error_messages=errors,
                    last_mode="add",
                    form_data=form
                )

            safe_name = secure_filename(uploaded.filename)
            path = os.path.join(UPLOAD_FOLDER, safe_name)
            uploaded.save(path)

            try:
                wb = load_workbook(path)
            except InvalidFileException:
                return render_template(
                    "index.html",
                    error_messages=["업로드한 파일이 엑셀 형식이 아닙니다. (.xlsx)"],
                    last_mode="add",
                    form_data=form
                )

            students = []
            for i in range(1, student_count + 1):
                name = form.get(f"add_student_{i}_name", "").strip()
                cat = form.get(f"add_student_{i}_category", "").strip()
                raw = form.get(f"add_student_{i}_answers", "")
                try:
                    s = int(form.get(f"add_student_{i}_start", 1))
                    e = int(form.get(f"add_student_{i}_end", 45))
                except ValueError:
                    s, e = None, None

                if not (name or raw):
                    continue

                parsed = parse_answers(raw)
                validate_range(f"{name}({cat})", s, e, len(parsed), errors)
                if cat not in wb.sheetnames:
                    errors.append(f"엑셀에 '{cat}' 시트가 없습니다.")
                students.append({"name": name, "category": cat, "answers": parsed, "start": s, "end": e})

            if errors:
                return render_template(
                    "index.html",
                    error_messages=errors,
                    last_mode="add",
                    form_data=form
                )

            # 학생 반영
            for s in students:
                ws = wb[s["category"]]
                append_or_update_student(ws, s["name"], s["answers"], s["start"], s["end"])
            wb.save(path)
            return send_file(path, as_attachment=True, download_name=safe_name)

        # mode 가 이상하면 그냥 처음 화면
        return render_template("index.html", last_mode="none", form_data=request.form)

    # GET 요청: 폼 초기 상태 (form_data를 빈 dict로 넘겨 오류 방지)
    return render_template("index.html", last_mode="none", form_data={})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
